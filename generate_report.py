"""
商品別売上CSV / 売上集計CSV / 月別売上目標.xlsx を読み込み、
直売所本店 売上レポートのフル版HTML（report_template.html）を生成する。

send_report.py の前段として実行される想定。
依存: pandas, openpyxl （Streamlit非依存）
"""

import calendar
import datetime
import glob
import os
import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

JST = datetime.timezone(datetime.timedelta(hours=9))


def get_report_date() -> datetime.date:
    """21:00-23:59は当日、00:00-20:59は前日をレポート対象日とする"""
    now = datetime.datetime.now(JST)
    if now.hour >= 21:
        return now.date()
    return now.date() - datetime.timedelta(days=1)

# ─── パス ───
BASE_DIR = Path(__file__).parent
TARGET_FILE = BASE_DIR / "月別売上目標.xlsx"
OUTPUT_HTML = BASE_DIR / "report_template.html"

# ─── カテゴリー ───
CATEGORIES = [
    "ソフトクリーム", "牛乳", "アイスクリーム", "ベーコン",
    "卵", "つくね", "餃子", "フード・弁当", "その他商品合計",
]
CATEGORY_DISPLAY = [
    "ソフトクリーム", "牛乳", "アイスクリーム", "ベーコン",
    "卵", "つくね・親鶏", "餃子", "フード・弁当", "その他",
]
WEEKDAYS_JA = ["月", "火", "水", "木", "金", "土", "日"]

PRODUCT_RULES = [
    ("フード・弁当", re.compile(r"カップミルク|ベーコン串|ソーセージ串|珈琲牛乳")),  # 牛乳・ベーコンルールより先に捕捉
    ("ソフトクリーム", re.compile(r"ソフトクリーム")),
    ("牛乳",         re.compile(r"牛乳")),
    ("アイスクリーム", re.compile(r"バニラ|アイス")),
    ("ベーコン",      re.compile(r"ベーコン|薫製|ソーセージ|ヴルスト|ウインナー")),
    ("卵",           re.compile(r"卵")),
    ("つくね",        re.compile(r"つくね|おやどり|親鶏")),
    ("餃子",         re.compile(r"餃子")),
    ("フード・弁当",   re.compile(r"ポップコーン|コーヒー|カフェオレ|弁当|フード|ハンバーグ|軽食|カレー|フライドポテト|パフェ|ホットドッグ|シフォンサンド|飲むヨーグルト")),
]
EXCLUDE_PATTERN = re.compile(
    r"送料|宅急便|クール便|レジ袋|保冷バッグ|資材|ワークショップ|レンタル|販売遊具|カスタム商品"
)


def classify_product(name: str):
    if EXCLUDE_PATTERN.search(name):
        return None
    for cat, pattern in PRODUCT_RULES:
        if pattern.search(name):
            return cat
    return "その他商品合計"


# ─── データ読み込み ───
def load_targets() -> pd.DataFrame:
    report_date = get_report_date()
    sheet_name = f"{report_date.month}月"
    wb = openpyxl.load_workbook(TARGET_FILE, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"[エラー] シート '{sheet_name}' が見つかりません: {TARGET_FILE}", file=sys.stderr)
        sys.exit(1)
    ws = wb[sheet_name]
    days_in_month = calendar.monthrange(report_date.year, report_date.month)[1]
    rows = []
    for row in ws.iter_rows(min_row=6, max_row=5 + days_in_month, values_only=True):
        if row[0] is None:
            continue
        dt = row[0]
        if isinstance(dt, datetime.datetime):
            dt = dt.date()
        rows.append({
            "日付": dt, "曜日": row[1], "類型": row[2], "売上指数": row[3],
            "ソフトクリーム": row[4], "牛乳": row[5], "アイスクリーム": row[6],
            "ベーコン": row[7], "卵": row[8], "つくね": row[9],
            "餃子": row[10], "フード・弁当": row[11], "その他商品合計": row[12],
            "日次合計": row[13],
        })
    return pd.DataFrame(rows)


def load_monthly_targets() -> list:
    """月別売上目標.xlsx の全シートから月次目標合計を取得（4月〜翌3月の年度順）"""
    wb = openpyxl.load_workbook(TARGET_FILE, data_only=True)
    report_date = get_report_date()
    fy_start_year = report_date.year if report_date.month >= 4 else report_date.year - 1

    month_order = [
        ("4月", 4), ("5月", 5), ("6月", 6), ("7月", 7),
        ("8月", 8), ("9月", 9), ("10月", 10), ("11月", 11),
        ("12月", 12), ("1月", 1), ("2月", 2), ("3月", 3),
    ]

    results = []
    for sheet_name, month_num in month_order:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        year = fy_start_year if month_num >= 4 else fy_start_year + 1
        total = 0.0
        for row in ws.iter_rows(min_row=6, values_only=True):
            if row[0] is None:
                break
            val = row[13]
            if val is not None:
                total += float(val)
        results.append({
            "sheet": sheet_name,
            "month_num": month_num,
            "year": year,
            "total": total,
            "first_date": datetime.date(year, month_num, 1),
        })
    return results


def load_monthly_actuals(daily_sales: pd.DataFrame) -> dict:
    """日別売上DataFrameを月別に集計して返す {date(year, month, 1): total}"""
    if daily_sales.empty:
        return {}
    df = daily_sales.copy()
    df["month_key"] = df["日付"].apply(lambda d: d.replace(day=1))
    return df.groupby("month_key")["売上"].sum().to_dict()


def load_daily_sales() -> pd.DataFrame:
    """日別売上(年月：*.csv) から日次売上合計を読み込む（最も正確なソース）"""
    files = sorted(glob.glob(str(BASE_DIR / "日別売上(年月*).csv")))
    if not files:
        files = sorted(glob.glob(str(BASE_DIR / "売上集計_*.csv")))
    if not files:
        return pd.DataFrame()

    dfs = []
    for f in files:
        try:
            df = pd.read_csv(f, encoding="cp932")
        except UnicodeDecodeError:
            df = pd.read_csv(f, encoding="utf-8")
        dfs.append(df)

    combined = pd.concat(dfs, ignore_index=True)
    col_date, col_sales = combined.columns[0], combined.columns[1]
    combined["日付"] = pd.to_datetime(combined[col_date], errors="coerce").dt.date
    combined["売上"] = pd.to_numeric(combined[col_sales], errors="coerce")
    result = combined[["日付", "売上"]].dropna()
    result = result[result["売上"] > 0]
    return result


def load_product_sales() -> pd.DataFrame:
    """単日の商品別売上CSVからカテゴリー別内訳を集計（複数日レンジは除外）"""
    patterns = [
        str(BASE_DIR / "商品別売上_*.csv"),
        str(BASE_DIR / "商品別売上(期間*).csv"),
    ]
    files = []
    for p in patterns:
        files.extend(glob.glob(p))
    files = sorted(set(files))
    if not files:
        return pd.DataFrame()

    records = []
    for f in files:
        basename = os.path.basename(f)
        m = re.search(r"(\d{8})-(\d{8})", basename)
        if not m:
            continue
        start_dt = datetime.datetime.strptime(m.group(1), "%Y%m%d").date()
        end_dt = datetime.datetime.strptime(m.group(2), "%Y%m%d").date()

        if start_dt != end_dt:
            continue

        try:
            df = pd.read_csv(f, encoding="cp932")
        except UnicodeDecodeError:
            df = pd.read_csv(f, encoding="utf-8")

        col_name = df.columns[1]
        col_sales = df.columns[3]

        for _, row in df.iterrows():
            name_raw = row[col_name]
            if pd.isna(name_raw):
                continue
            name = str(name_raw)
            cat = classify_product(name)
            if cat is None:
                continue
            sales = pd.to_numeric(row[col_sales], errors="coerce")
            if pd.isna(sales) or sales <= 0:
                continue
            records.append({"日付": start_dt, "カテゴリー": cat, "売上": sales})

    if not records:
        return pd.DataFrame()
    rdf = pd.DataFrame(records)
    pivot = rdf.groupby(["日付", "カテゴリー"])["売上"].sum().reset_index()
    result = pivot.pivot(index="日付", columns="カテゴリー", values="売上").fillna(0)
    return result.reset_index()


# ─── HTMLヘルパー ───
def color_class(pct: float) -> str:
    if pct >= 100:
        return "over"
    if pct >= 85:
        return "near"
    return "under"


def fmt(val) -> str:
    return f"{val:,.0f}"


def bar(pct: float, h: int = 8) -> str:
    cc = color_class(pct)
    w = min(pct / 120 * 100, 100)
    return (
        f'<div class="bar-wrap" style="height:{h}px;">'
        f'<div class="bar-fill b-{cc}" style="width:{w:.1f}%;height:{h}px;"></div>'
        f'<div class="bar-target" style="top:-3px;height:{h + 6}px;"></div>'
        f'</div>'
    )


def scale_html() -> str:
    return (
        '<div class="bar-scale-mid">'
        '<span class="l0">0</span><span class="l100">100%</span><span class="l120">120%</span>'
        '</div>'
    )


CSS = """
* { box-sizing: border-box; margin: 0; padding: 0; }
body {
  font-family: -apple-system, 'Hiragino Sans', 'Meiryo', sans-serif;
  background: #e8eaed;
  display: flex;
  justify-content: center;
  padding: 16px;
}
.card {
  width: 360px;
  background: #fff;
  border-radius: 12px;
  overflow: hidden;
  box-shadow: 0 2px 12px rgba(0,0,0,0.15);
  transition: width 0.25s ease;
}
body.pc-view .card { width: min(900px, 96vw); }
.header {
  background: #1c3f60;
  color: #fff;
  padding: 10px 14px 9px;
  display: flex;
  justify-content: space-between;
  align-items: flex-start;
}
.header .title { font-size: 15px; font-weight: bold; }
.header .period { font-size: 11px; color: #7db3d6; margin-top: 4px; }
.header-right { display: flex; flex-direction: column; align-items: flex-end; gap: 5px; }
.view-toggle {
  background: rgba(255,255,255,0.15);
  border: 1px solid rgba(255,255,255,0.3);
  color: #fff;
  font-size: 10px;
  padding: 3px 9px;
  border-radius: 4px;
  cursor: pointer;
  font-family: inherit;
  white-space: nowrap;
}
.view-toggle:hover { background: rgba(255,255,255,0.25); }
.pc-grid { }
body.pc-view .pc-grid { display: grid; grid-template-columns: 1fr 1fr; }
body.pc-view .pc-left { border-right: 1px solid #efefef; }
.section {
  padding: 10px 14px 11px;
  border-bottom: 1px solid #efefef;
}
.section:last-of-type { border-bottom: none; }
.section-title {
  font-size: 10px;
  font-weight: bold;
  color: #999;
  letter-spacing: 0.06em;
  text-transform: uppercase;
  margin-bottom: 8px;
}
.c-over  { color: #2a9658; }
.c-near  { color: #d4860f; }
.c-under { color: #c94040; }
.b-over  { background: #2a9658; }
.b-near  { background: #d4860f; }
.b-under { background: #c94040; }
.bar-wrap {
  position: relative;
  background: #f0f0f0;
  border-radius: 3px;
  height: 8px;
}
.bar-fill { position: absolute; top: 0; left: 0; border-radius: 3px; }
.bar-target {
  position: absolute;
  left: calc(83.33% - 1px);
  width: 2px;
  background: #aaa;
  border-radius: 1px;
}
.bar-scale-mid { position: relative; height: 10px; margin-top: 2px; }
.bar-scale-mid .l0   { position:absolute; left:0;      font-size:8px; color:#ccc; }
.bar-scale-mid .l100 { position:absolute; left:83.33%; font-size:8px; color:#bbb; transform:translateX(-50%); }
.bar-scale-mid .l120 { position:absolute; right:0;     font-size:8px; color:#ccc; }
.day-list { display: flex; flex-direction: column; gap: 7px; }
.day-row-top { display: flex; align-items: baseline; gap: 0; margin-bottom: 3px; }
.day-label { font-size: 12px; font-weight: 500; color: #333; width: 54px; flex-shrink: 0; }
.day-amounts {
  flex: 1; font-size: 10px; color: #999;
  white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
}
.day-amounts .actual { color: #444; font-weight: 500; }
.pct-sm { font-size: 13px; font-weight: bold; width: 48px; text-align: right; flex-shrink: 0; }
.day-highlight { background: #edf2f7; border-radius: 6px; padding: 5px 7px; margin: 0 -7px; }
.summary-row { display: flex; align-items: flex-end; gap: 10px; margin-bottom: 7px; }
.pct-lg { font-size: 20px; font-weight: bold; line-height: 1; flex-shrink: 0; width: 68px; }
.summary-detail { flex: 1; font-size: 11px; color: #555; line-height: 1.7; padding-bottom: 1px; }
.summary-detail .amt { color: #333; font-weight: 500; }
.summary-detail .sub { font-size: 10px; color: #999; }
.cat-list { display: flex; flex-direction: column; gap: 7px; }
.cat-row-top { display: flex; align-items: baseline; margin-bottom: 3px; gap: 6px; }
.cat-name { font-size: 12px; font-weight: 500; color: #222; width: 90px; flex-shrink: 0; white-space: nowrap; }
.cat-amounts {
  flex: 1; font-size: 10px; color: #999;
  white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
}
.cat-amounts .actual { color: #444; font-weight: 500; }
.abs-wrap { position: relative; background: #f0f0f0; border-radius: 3px; height: 8px; }
.abs-fill { position: absolute; top: 0; left: 0; height: 8px; border-radius: 3px; }
.abs-target { position: absolute; top: -3px; width: 2px; height: 14px; background: #888; border-radius: 1px; }
.abs-scale { display: flex; justify-content: space-between; font-size: 8px; color: #ccc; margin-top: 3px; }
.month-list { display: flex; flex-direction: column; gap: 6px; }
.month-row { display: flex; align-items: center; gap: 6px; }
.month-label-col { font-size: 11px; color: #555; font-weight: 500; width: 28px; flex-shrink: 0; text-align: right; }
.month-current-row .month-label-col { color: #1c3f60; font-weight: 700; }
.month-bar-area { flex: 1; }
.month-track { position: relative; background: #f0f0f0; border-radius: 3px; height: 14px; overflow: hidden; }
.month-current-row .month-track { background: #edf2f7; }
.month-future-track { background: #f8f8f8; border: 1px dashed #ddd; overflow: visible; }
.month-actual { position: absolute; top: 0; left: 0; height: 14px; border-radius: 3px; }
.month-target-line { position: absolute; top: 0; width: 2px; height: 14px; background: rgba(80,80,80,0.5); }
.month-sub { font-size: 9px; color: #aaa; margin-top: 2px; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
.month-pct-col { font-size: 11px; font-weight: bold; width: 44px; flex-shrink: 0; text-align: right; }
body.pc-view .month-track { height: 18px; }
body.pc-view .month-actual { height: 18px; }
body.pc-view .month-target-line { height: 18px; }
body.pc-view .month-list { gap: 8px; }
.no-data { font-size: 11px; color: #bbb; text-align: center; padding: 8px 0; }
.footer {
  padding: 6px 14px 10px; font-size: 9px; color: #bbb;
  display: flex; justify-content: space-between;
}
"""

JS = """
<script>
function toggleView() {
  var isPC = document.body.classList.toggle('pc-view');
  document.querySelector('.view-toggle').textContent = isPC ? '📱 スマホ表示' : '🖥 PC表示';
}
</script>
"""


def build_summary(targets: pd.DataFrame, report_date: datetime.date = None) -> dict:
    """本日/週間/月間の実績・目標・達成率を dict で返す。
    send_report.py のテキスト本文生成にも使う。"""
    if report_date is None:
        report_date = get_report_date()
    available = targets[(targets["実績"].notna()) & (targets["日付"] <= report_date)]
    if available.empty:
        raise RuntimeError("実績データが目標期間内に見つかりません")
    if report_date not in available["日付"].values:
        report_date = available["日付"].max()
        print(f"[警告] 当日データなし。最新データ日付 ({report_date}) を使用")
    week_start = report_date - datetime.timedelta(days=report_date.weekday())

    today_row = available[available["日付"] == report_date].iloc[0]
    week_data = targets[
        (targets["日付"] >= week_start)
        & (targets["日付"] <= report_date)
        & (targets["実績"].notna())
    ]
    month_data = targets[
        (targets["日付"] >= report_date.replace(day=1))
        & (targets["日付"] <= report_date)
        & (targets["実績"].notna())
    ]

    def block(actual, target):
        pct = actual / target * 100 if target > 0 else 0
        return {"actual": float(actual), "target": float(target), "pct": pct}

    return {
        "report_date": report_date,
        "weekday": WEEKDAYS_JA[report_date.weekday()],
        "today": block(today_row["実績"], today_row["日次合計"]),
        "week": {
            **block(week_data["実績"].sum(), week_data["日次合計"].sum()),
            "start": week_start if not week_data.empty else report_date,
            "end": report_date,
        },
        "month": {
            **block(month_data["実績"].sum(), month_data["日次合計"].sum()),
            "days_elapsed": len(month_data),
            "days_in_month": calendar.monthrange(report_date.year, report_date.month)[1],
        },
    }


def build_monthly_chart_html(
    monthly_targets: list,
    monthly_actuals: dict,
    report_date: datetime.date,
) -> str:
    """月別売上グラフセクションのHTMLを返す"""
    if not monthly_targets:
        return ""

    scale_max = max(m["total"] for m in monthly_targets) * 1.2
    if scale_max == 0:
        return ""

    current_month_key = report_date.replace(day=1)
    fy_start_year = monthly_targets[0]["year"] if monthly_targets[0]["month_num"] >= 4 else monthly_targets[0]["year"] - 1
    fy_end_year = fy_start_year + 1

    h = [
        '<div class="section">',
        f'<div class="section-title">月別売上実績 vs 目標（{fy_start_year}年4月〜{fy_end_year}年3月）</div>',
        '<div class="month-list">',
    ]

    for m in monthly_targets:
        month_key = m["first_date"]
        total_target = m["total"]
        target_pct = total_target / scale_max * 100
        target_man = round(total_target / 10000)

        if month_key < current_month_key:
            # 過去月: 実績バー + 目標線 + 達成率%
            actual = monthly_actuals.get(month_key, 0)
            actual_pct = min(actual / scale_max * 100, 100)
            pct = actual / total_target * 100 if total_target > 0 else 0
            cc = color_class(pct)
            actual_man = round(actual / 10000)
            h.append(
                f'<div class="month-row">'
                f'<div class="month-label-col">{m["sheet"]}</div>'
                f'<div class="month-bar-area">'
                f'<div class="month-track">'
                f'<div class="month-actual b-{cc}" style="width:{actual_pct:.1f}%;"></div>'
                f'<div class="month-target-line" style="left:calc({target_pct:.1f}% - 1px);"></div>'
                f'</div>'
                f'<div class="month-sub">実績 {actual_man}万 / 目標 {target_man}万</div>'
                f'</div>'
                f'<div class="month-pct-col c-{cc}">{pct:.0f}%</div>'
                f'</div>'
            )

        elif month_key == current_month_key:
            # 当月: 途中実績バー + 進行中表示
            actual = monthly_actuals.get(month_key, 0)
            actual_pct = min(actual / scale_max * 100, 100)
            days_elapsed = report_date.day
            days_in_month = calendar.monthrange(report_date.year, report_date.month)[1]
            actual_man = round(actual / 10000)
            pct_vs_target = actual / total_target * 100 if total_target > 0 else 0
            h.append(
                f'<div class="month-row month-current-row">'
                f'<div class="month-label-col">{m["sheet"]}</div>'
                f'<div class="month-bar-area">'
                f'<div class="month-track">'
                f'<div class="month-actual b-under" style="width:{actual_pct:.1f}%;opacity:0.75;"></div>'
                f'<div class="month-target-line" style="left:calc({target_pct:.1f}% - 1px);"></div>'
                f'</div>'
                f'<div class="month-sub" style="color:#7da0c0;">'
                f'進行中 実績 {actual_man}万 / 目標 {target_man}万（{days_elapsed}/{days_in_month}日）'
                f'</div>'
                f'</div>'
                f'<div class="month-pct-col c-under" style="line-height:1.3;">'
                f'<span style="font-size:9px;font-weight:normal;color:#aaa;">月次目標比</span><br>'
                f'{pct_vs_target:.0f}%'
                f'</div>'
                f'</div>'
            )

        else:
            # 未来月: 目標線のみ
            h.append(
                f'<div class="month-row">'
                f'<div class="month-label-col" style="color:#ccc;">{m["sheet"]}</div>'
                f'<div class="month-bar-area">'
                f'<div class="month-track month-future-track">'
                f'<div class="month-target-line" style="left:calc({target_pct:.1f}% - 1px);background:rgba(180,180,180,0.5);"></div>'
                f'</div>'
                f'<div class="month-sub">目標 {target_man}万</div>'
                f'</div>'
                f'<div class="month-pct-col" style="color:#ccc;">—</div>'
                f'</div>'
            )

    scale_max_man = round(scale_max / 10000)
    h.append(
        '</div>'
        '<div style="display:flex;justify-content:space-between;font-size:8px;color:#ccc;'
        'margin-top:6px;padding-top:4px;border-top:1px solid #f5f5f5;">'
        f'<span>0</span><span>縦線 = 月次目標　バー = 実績</span><span>{scale_max_man}万（最大）</span>'
        '</div>'
        '</div>'
    )
    return "".join(h)


def build_html(
    targets: pd.DataFrame,
    has_product_data: bool,
    report_date: datetime.date = None,
    monthly_targets: list = None,
    monthly_actuals: dict = None,
) -> str:
    if report_date is None:
        report_date = get_report_date()
    available = targets[(targets["実績"].notna()) & (targets["日付"] <= report_date)]
    if available.empty:
        raise RuntimeError("実績データが目標期間内に見つかりません")
    if report_date not in available["日付"].values:
        report_date = available["日付"].max()
    report_wd = WEEKDAYS_JA[report_date.weekday()]

    week_start = report_date - datetime.timedelta(days=report_date.weekday())
    week_data = targets[
        (targets["日付"] >= week_start)
        & (targets["日付"] <= report_date)
        & (targets["実績"].notna())
    ]
    month_data = targets[
        (targets["日付"] >= report_date.replace(day=1))
        & (targets["日付"] <= report_date)
        & (targets["実績"].notna())
    ]

    h = [
        '<!DOCTYPE html><html lang="ja"><head><meta charset="UTF-8">',
        '<meta name="viewport" content="width=device-width,initial-scale=1.0">',
        f'<style>{CSS}</style></head><body>',
        '<div class="card">',
        # ヘッダー（PCトグルボタン付き）
        '<div class="header">'
        '<div>'
        f'<div class="title">直売所本店 売上レポート</div>'
        f'<div class="period">{report_date.strftime("%Y/%m/%d")}（{report_wd}）</div>'
        '</div>'
        '<div class="header-right">'
        '<button class="view-toggle" onclick="toggleView()">🖥 PC表示</button>'
        '</div>'
        '</div>',
        # PC 2カラムグリッド開始（左:① 右:③）
        '<div class="pc-grid"><div class="pc-left">',
    ]

    # ① 日次達成率
    h.append('<div class="section"><div class="section-title">日次達成率（売上全体）</div><div class="day-list">')
    for _, row in week_data.iterrows():
        dt = row["日付"]
        wd = WEEKDAYS_JA[dt.weekday()]
        actual, target = row["実績"], row["日次合計"]
        pct = actual / target * 100 if target > 0 else 0
        cc = color_class(pct)
        is_last = (dt == report_date)

        if is_last:
            h.append('<div class="day-highlight">')
            h.append(
                f'<div class="day-row-top">'
                f'<div class="day-label" style="color:#1c3f60;font-weight:700;width:auto;margin-right:4px;">'
                f'{dt.month}/{dt.day} {wd}<span style="font-size:9px;font-weight:500;color:#7da0c0;margin-left:3px;">本日</span></div>'
                f'<div class="day-amounts"><span class="actual">{fmt(actual)}</span> / {fmt(target)}円</div>'
                f'<div class="pct-sm c-{cc}">{pct:.0f}%</div></div>'
            )
        else:
            h.append('<div>')
            h.append(
                f'<div class="day-row-top">'
                f'<div class="day-label">{dt.month}/{dt.day} {wd}</div>'
                f'<div class="day-amounts"><span class="actual">{fmt(actual)}</span> / {fmt(target)}円</div>'
                f'<div class="pct-sm c-{cc}">{pct:.0f}%</div></div>'
            )
        h.append(bar(pct))
        h.append('</div>')
    h.append('</div></div>')  # /day-list /section①

    # pc-left終了 → pc-right開始（③ カテゴリー別）
    h.append('</div><div class="pc-right">')

    # ③ カテゴリー別（週間・絶対スケール）
    h.append('<div class="section"><div class="section-title">週間累計（カテゴリー別）</div>')
    if has_product_data:
        cat_week_targets = [week_data[cat].sum() for cat in CATEGORIES]
        max_target = max(cat_week_targets) if any(t > 0 for t in cat_week_targets) else 1
        abs_max = max_target * 1.2

        h.append('<div class="cat-list">')
        for cat, disp in zip(CATEGORIES, CATEGORY_DISPLAY):
            col = f"{cat}_実績"
            ca = week_data[col].sum() if col in week_data.columns else 0
            ct = week_data[cat].sum()
            cp = ca / ct * 100 if ct > 0 else 0
            cc = color_class(cp)
            target_pct = ct / abs_max * 100
            actual_pct = min(ca / abs_max * 100, 100)
            h.append(
                f'<div><div class="cat-row-top">'
                f'<div class="cat-name">{disp}</div>'
                f'<div class="cat-amounts"><span class="actual">{fmt(ca)}</span> / {fmt(ct)}円</div>'
                f'<div class="pct-sm c-{cc}">{cp:.0f}%</div></div>'
                f'<div class="abs-wrap">'
                f'<div class="abs-fill b-{cc}" style="width:{actual_pct:.1f}%;"></div>'
                f'<div class="abs-target" style="left:calc({target_pct:.1f}% - 1px);"></div>'
                f'</div></div>'
            )
        h.append('</div>')
        max_man = round(max_target / 10000)
        max_man_120 = round(max_target * 1.2 / 10000)
        h.append(
            f'<div class="abs-scale" style="margin-top:5px;">'
            f'<span>0</span>'
            f'<span>{max_man}万（最大目標）</span>'
            f'<span>{max_man_120}万+</span>'
            f'</div>'
        )
    else:
        h.append('<div class="no-data">商品別売上データなし（売上集計CSVのみ）</div>')
    h.append('</div>')  # /section③

    # pc-grid終了
    h.append('</div></div>')  # /pc-right /pc-grid

    # ② 週間累計（全商品合計）- 全幅
    wa, wt = week_data["実績"].sum(), week_data["日次合計"].sum()
    wp = wa / wt * 100 if wt > 0 else 0
    wc = color_class(wp)
    h.append(
        f'<div class="section"><div class="section-title">週間累計（全商品合計）</div>'
        f'<div class="summary-row">'
        f'<div class="pct-lg c-{wc}">{wp:.0f}%</div>'
        f'<div class="summary-detail">'
        f'<span class="amt">実績 {fmt(wa)}円</span> / <span>目標 {fmt(wt)}円</span>'
        f'</div></div>'
        f'{bar(wp)}{scale_html()}</div>'
    )

    # ④ 月間累計 - 全幅
    ma, mt = month_data["実績"].sum(), month_data["日次合計"].sum()
    mp = ma / mt * 100 if mt > 0 else 0
    mc = color_class(mp)
    de = len(month_data)
    days_in_month = calendar.monthrange(report_date.year, report_date.month)[1]
    dr = days_in_month - de
    h.append(
        f'<div class="section">'
        f'<div class="section-title">月間累計（{report_date.month}/1〜{report_date.month}/{report_date.day}）</div>'
        f'<div class="summary-row">'
        f'<div class="pct-lg c-{mc}">{mp:.0f}%</div>'
        f'<div class="summary-detail">'
        f'<span class="amt">実績 {fmt(ma)}円</span> / <span>目標 {fmt(mt)}円</span><br>'
        f'<span class="sub">{de}日経過 / {days_in_month}日　残{dr}日</span>'
        f'</div></div>'
        f'{bar(mp)}{scale_html()}</div>'
    )

    # ⑤ 月別グラフ - 全幅
    if monthly_targets:
        h.append(build_monthly_chart_html(monthly_targets, monthly_actuals or {}, report_date))

    # フッター
    src = "商品別売上CSV" if has_product_data else "売上集計CSV"
    h.append(
        f'<div class="footer">'
        f'<span>縦線 = 目標100%　グラフ最大 = 120%</span>'
        f'<span>データ元: {src}</span></div>'
    )

    h.append('</div>')  # /card
    h.append(JS)
    h.append('</body></html>')
    return "".join(h)


def main() -> Path:
    if not TARGET_FILE.exists():
        print(f"[エラー] 目標ファイルが見つかりません: {TARGET_FILE}", file=sys.stderr)
        sys.exit(1)

    targets = load_targets()
    if targets.empty:
        print("[エラー] 月別売上目標.xlsx の読み込みに失敗", file=sys.stderr)
        sys.exit(1)

    daily_sales = load_daily_sales()
    if daily_sales.empty:
        print("[エラー] 実績データが見つかりません", file=sys.stderr)
        sys.exit(1)
    targets = targets.merge(
        daily_sales.rename(columns={"売上": "実績"}), on="日付", how="left",
    )

    product_sales = load_product_sales()
    has_product_data = not product_sales.empty
    if has_product_data:
        for cat in CATEGORIES:
            if cat in product_sales.columns:
                col_map = product_sales[["日付", cat]].rename(columns={cat: f"{cat}_実績"})
                targets = targets.merge(col_map, on="日付", how="left")
            else:
                targets[f"{cat}_実績"] = 0.0

    monthly_targets = load_monthly_targets()
    monthly_actuals = load_monthly_actuals(daily_sales)

    report_date = get_report_date()
    html = build_html(targets, has_product_data, report_date, monthly_targets, monthly_actuals)
    OUTPUT_HTML.write_text(html, encoding="utf-8")
    summary = build_summary(targets, report_date)
    print(f"[OK] HTML生成: {OUTPUT_HTML}  ({len(html):,} bytes)")
    return OUTPUT_HTML, summary, targets, has_product_data


if __name__ == "__main__":
    main()
